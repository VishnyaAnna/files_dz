import zipfile, os
from os.path import basename
from openpyxl import load_workbook
import csv

#путь к папкам
path_inputfiles = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
path_outputfiles = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')


files_dir = os.listdir(path_inputfiles)  # список имен фалов в папке "files"
path_zip = os.path.join(path_outputfiles, "test.zip")  # путь к архиву

# добавление zip файла
with zipfile.ZipFile(path_zip, mode='w',
                     compression=zipfile.ZIP_DEFLATED) as zf:
    for file in files_dir:
        add_file = os.path.join(path_inputfiles, file)
        zf.write(add_file, basename(add_file))

# Проверка содержимого xlsx
with zipfile.ZipFile(path_zip) as zf:
    xf = zf.extract("file_xlsx.xlsx")
    xlsxfile = load_workbook(xf)
    sheet = xlsxfile.active
    assert sheet.cell(row=7, column=1).value == "February", f"Expected result: {'February'}, " \
                                                           f"actual result: {sheet.cell(row=7, column=1).value}"
    os.remove("file_xlsx.xlsx")

# Проверка содержимого pdf
with zipfile.ZipFile(path_zip, mode='r') as zip_file:
    text = str(zip_file.read('file_pdf.pdf'))
    assert text.__contains__('PDF')

# Проверка содержимого csv
with zipfile.ZipFile(path_zip) as zf:
    cf = zf.extract("file_csv.csv")
    with open(cf) as csvfile:
        csvfile = csv.reader(csvfile)
        list_csv = []
        for r in csvfile:
            text = "".join(r).replace(";", " ")
            list_csv.append(text)

        assert list_csv[
                   0] == "Anna Pavel Peter", f"Expected result: {'Anna Pavel Peter'}, " \
                                                        f"actual result: {list_csv[0]}"
        os.remove("file_csv.csv")

