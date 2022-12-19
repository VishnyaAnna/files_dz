import zipfile, os
from os.path import basename
from openpyxl import load_workbook
import csv
from PyPDF2 import PdfReader

# путь к папкам
path_inputfiles = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources/files')
path_outputfiles = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')

files_dir = os.listdir(path_inputfiles)  # список имен фалов в папке "files"
path_zip = os.path.join(path_outputfiles, "test.zip")  # путь к архиву

# добавление zip файла
with zipfile.ZipFile(path_zip, mode='w',
                     compression=zipfile.ZIP_DEFLATED) as zf:
    for file in files_dir:
        add_file = os.path.join(path_inputfiles, file)
        zf.write(add_file, basename(add_file))


# Проверка содержимого xlsx
def test_check_xlsx_file():
    with zipfile.ZipFile(path_zip) as zf:
        xf = zf.extract("file_xlsx.xlsx")
        xlsxfile = load_workbook(xf)
        sheet = xlsxfile.active
        assert sheet.cell(row=7, column=1).value == "February", f"Expected result: {'February'}, " \
                                                                f"actual result: {sheet.cell(row=7, column=1).value}"
        os.remove("file_xlsx.xlsx")


# Проверка содержимого pdf
def test_check_pdf_file():
    with zipfile.ZipFile(path_zip) as zip_file:
        file_pdf = zip_file.extract('file_pdf.pdf')
        reader = PdfReader(file_pdf)
        page = reader.pages[0]
        text = str(page.extract_text())
        assert text.__contains__('PDF')

    os.remove("file_pdf.pdf")


# Проверка содержимого csv
def test_check_csv_file():
    with zipfile.ZipFile(path_zip) as zf:
        cf = zf.extract("file_csv.csv")
        with open(cf) as csvfile:
            csvfile = csv.reader(csvfile)
            list_csv = []
            for r in csvfile:
                text = "".join(r).replace(";", " ")
                list_csv.append(text)

                assert list_csv[0] == "Anna Pavel Peter", \
                    f"Expected result: {'Anna Pavel Peter'}, actual result: {list_csv[0]}"

        os.remove("file_csv.csv")
